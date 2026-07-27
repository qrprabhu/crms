from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import models, transaction
from django.db.models import Q
from django.utils import timezone

from accounts.models import Account
from contacts.models import Contact
from deals.models import Deal
from inventory.models import Product, Vendor
from leads.models import Lead

from .models import (
    SupportActivity,
    SupportAttachment,
    SupportCase,
    SupportComment,
    SupportEmailLog,
    SupportImportJob,
    SupportLinkedRecord,
    SupportNote,
    SupportSolution,
    SupportTimelineEntry,
)


CASE_IMPORT_FIELD_MAP = {
    "case_id": "case_number",
    "case_number": "case_number",
    "subject": "subject",
    "status": "status",
    "priority": "priority",
    "case_origin": "case_origin",
    "case_reason": "case_reason",
    "type": "type",
    "description": "description",
    "internal_comments": "internal_comments",
    "solution": "solution_text",
    "solution_text": "solution_text",
    "reported_by": "reported_by",
    "email": "email",
    "company": "company",
    "country": "country",
    "phone": "phone",
    "lead": "lead",
    "lead_name": "lead_name",
    "lead_source": "lead_source",
    "case_owner": "owner",
    "owner": "owner",
    "product_name": "product",
    "product": "product",
    "related_to": "related_contact",
    "related_contact": "related_contact",
    "account_name": "account",
    "account": "account",
    "deal_name": "deal",
    "deal": "deal",
}

SOLUTION_IMPORT_FIELD_MAP = {
    "solution_id": "solution_number",
    "solution_number": "solution_number",
    "solution_title": "solution_title",
    "status": "status",
    "question": "question",
    "answer": "answer",
    "solution_owner": "owner",
    "owner": "owner",
    "product_name": "product",
    "product": "product",
    "connected_to": "connected_to",
}


def _parse_uploaded_file(file_field) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = Path(file_field.name).suffix.lower()
    if suffix == ".csv":
        raw = file_field.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames or []), rows

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DjangoValidationError("openpyxl is required to parse XLSX files.") from exc

        workbook = load_workbook(file_field, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
        rows: list[dict[str, Any]] = []
        for row in values[1:]:
            if row is None:
                continue
            item = {}
            is_empty = True
            for index, header in enumerate(headers):
                value = row[index] if index < len(row) else None
                if value not in (None, ""):
                    is_empty = False
                item[header] = value
            if not is_empty:
                rows.append(item)
        return headers, rows

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise DjangoValidationError("xlrd is required to parse XLS files.") from exc

        workbook = xlrd.open_workbook(file_contents=file_field.read())
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return [], []
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        rows = []
        for row_idx in range(1, sheet.nrows):
            item = {}
            is_empty = True
            for col_idx, header in enumerate(headers):
                value = sheet.cell_value(row_idx, col_idx)
                if value not in ("", None):
                    is_empty = False
                item[header] = value
            if not is_empty:
                rows.append(item)
        return headers, rows

    raise DjangoValidationError("Unsupported file type. Supported formats: CSV, XLS, XLSX.")


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _resolve_user(value: Any):
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return UserProxy.get(value)
    text = str(value).strip()
    if text.isdigit():
        return UserProxy.get(int(text))
    return UserProxy.get_by_email(text)


class UserProxy:
    @staticmethod
    def get(pk: int):
        from django.contrib.auth import get_user_model

        return get_user_model().objects.filter(pk=pk, is_active=True).first()

    @staticmethod
    def get_by_email(email: str):
        from django.contrib.auth import get_user_model

        return get_user_model().objects.filter(email__iexact=email, is_active=True).first()


def _resolve_fk(field_name: str, value: Any):
    if value in (None, ""):
        return None
    if field_name == "owner":
        return _resolve_user(value)
    if field_name == "product":
        if isinstance(value, int) or str(value).isdigit():
            return Product.objects.filter(pk=int(value), is_active=True).first()
        return Product.objects.filter(Q(product_name__iexact=str(value).strip()) | Q(product_code__iexact=str(value).strip()), is_active=True).first()
    if field_name == "related_contact":
        if isinstance(value, int) or str(value).isdigit():
            return Contact.objects.filter(pk=int(value), is_active=True).first()
        text = str(value).strip()
        return Contact.objects.filter(
            Q(first_name__iexact=text) | Q(last_name__iexact=text) | Q(email__iexact=text),
            is_active=True,
        ).first()
    if field_name == "account":
        if isinstance(value, int) or str(value).isdigit():
            return Account.objects.filter(pk=int(value), is_active=True).first()
        return Account.objects.filter(account_name__iexact=str(value).strip(), is_active=True).first()
    if field_name == "deal":
        if isinstance(value, int) or str(value).isdigit():
            return Deal.objects.filter(pk=int(value), is_active=True).first()
        return Deal.objects.filter(deal_name__iexact=str(value).strip(), is_active=True).first()
    if field_name == "lead":
        if isinstance(value, int) or str(value).isdigit():
            return Lead.objects.filter(pk=int(value)).first()
        text = str(value).strip()
        return Lead.objects.filter(
            Q(first_name__iexact=text) | Q(last_name__iexact=text) | Q(email__iexact=text),
        ).first()
    return value


def log_timeline(module_type: str, record_id: int, action_type: str, message: str, user=None, metadata: dict | None = None):
    entry = SupportTimelineEntry.objects.create(
        module_type=module_type,
        record_id=record_id,
        action_type=action_type,
        message=message,
        created_by=user,
        metadata=metadata or {},
    )
    if module_type == SupportTimelineEntry.ModuleType.CASE:
        SupportCase.objects.filter(pk=record_id).update(last_activity_at=entry.created_at)
    if module_type == SupportTimelineEntry.ModuleType.SOLUTION:
        SupportSolution.objects.filter(pk=record_id).update(last_activity_at=entry.created_at)
    return entry


def list_timeline(module_type: str, record_id: int):
    return SupportTimelineEntry.objects.filter(module_type=module_type, record_id=record_id).select_related("created_by")


def list_cases():
    return SupportCase.objects.filter(is_active=True).select_related(
        "owner",
        "product",
        "related_contact",
        "account",
        "deal",
        "created_by",
        "updated_by",
    )


def list_solutions():
    return SupportSolution.objects.filter(is_active=True).select_related(
        "owner",
        "product",
        "source_case",
        "created_by",
        "updated_by",
    )


def get_case(case_id: int):
    return list_cases().get(pk=case_id)


def get_solution(solution_id: int):
    return list_solutions().get(pk=solution_id)


def _first_active_deal_for_contact(contact: Contact | None):
    if not contact:
        return None
    return (
        Deal.objects.filter(is_active=True, contact=contact)
        .select_related("account", "contact")
        .order_by("-created_at")
        .first()
    )


def _default_case_reason(case_type: str | None) -> str:
    mapping = {
        "complaint": "Product Issue",
        "problem": "Product Issue",
        "question": "Product Usage",
        "feature request": "Feature Request",
    }
    normalized = str(case_type or "").strip().lower()
    return mapping.get(normalized, "General Inquiry")


def _case_subject_for(payload: dict[str, Any]) -> str:
    product_name = getattr(payload.get("product"), "product_name", "") or "Support"
    contact_name = ""
    contact = payload.get("related_contact")
    if contact:
        contact_name = f"{contact.first_name} {contact.last_name}".strip()
    return f"{product_name} Issue - {contact_name}".strip(" -")


def _solution_title_for(payload: dict[str, Any]) -> str:
    source_case = payload.get("source_case")
    product_name = getattr(payload.get("product"), "product_name", "")
    if source_case and source_case.subject:
        return f"{source_case.subject} - Solution"
    if product_name:
        return f"{product_name} - Issue Resolution"
    return "Issue Resolution"


def _sync_support_links(module_type: str, obj, *, contact=None, account=None, deal=None, product=None, case=None, lead=None):
    SupportLinkedRecord.objects.filter(**_record_kwargs(module_type, obj)).delete()
    link_payload = {
        **_record_kwargs(module_type, obj),
        "contact": contact,
        "account": account,
        "deal": deal,
        "product": product,
        "lead": lead,
        "relationship_label": "Connected To",
        "metadata": {"autolink": True},
    }
    if module_type == "solution" and case:
        link_payload["case"] = case
    SupportLinkedRecord.objects.create(**link_payload)


@transaction.atomic
def create_case(data: dict[str, Any], user):
    payload = data.copy()
    payload.setdefault("owner", user)
    payload.setdefault("case_origin", "Web")
    payload.setdefault("status", "Open")
    payload.setdefault("priority", "Medium")
    payload.setdefault("case_reason", _default_case_reason(payload.get("type")))
    lead = payload.get("lead")
    if lead:
        payload.setdefault("lead_name", f"{lead.first_name} {lead.last_name}".strip())
        payload.setdefault("lead_source", lead.lead_source)
        payload.setdefault("company", payload.get("company") or lead.company)
        payload.setdefault("email", payload.get("email") or lead.email)
        payload.setdefault("phone", payload.get("phone") or lead.phone or lead.mobile)
        payload.setdefault("reported_by", payload.get("reported_by") or f"{lead.first_name} {lead.last_name}".strip())
    contact = payload.get("related_contact")
    if contact:
        payload.setdefault("account", contact.account)
        payload.setdefault("email", contact.email)
        payload.setdefault("phone", contact.phone)
        payload.setdefault("reported_by", f"{contact.first_name} {contact.last_name}".strip())
        if contact.account:
            payload.setdefault("company", contact.account.account_name)
        inferred_deal = payload.get("deal") or _first_active_deal_for_contact(contact)
        if inferred_deal:
            payload.setdefault("deal", inferred_deal)
    deal = payload.get("deal")
    if deal:
        payload.setdefault("account", deal.account)
        payload.setdefault("related_contact", deal.contact or payload.get("related_contact"))
    payload.setdefault("subject", _case_subject_for(payload))
    payload["created_by"] = user
    payload["updated_by"] = user
    case = SupportCase.objects.create(**payload)
    _sync_support_links(
        "case",
        case,
        contact=case.related_contact,
        account=case.account,
        deal=case.deal,
        product=case.product,
        lead=case.lead,
    )
    log_timeline("case", case.pk, "created", "Case Created", user=user, metadata={"case_number": case.case_number})
    return case


@transaction.atomic
def update_case(case: SupportCase, data: dict[str, Any], user):
    changed_fields = []
    for field, value in data.items():
        if getattr(case, field) != value:
            setattr(case, field, value)
            changed_fields.append(field)
    if case.related_contact and not case.account:
        case.account = case.related_contact.account
        changed_fields.append("account")
    if case.related_contact:
        case.email = case.email or case.related_contact.email
        case.phone = case.phone or case.related_contact.phone
        case.reported_by = case.reported_by or f"{case.related_contact.first_name} {case.related_contact.last_name}".strip()
    if case.related_contact and not case.deal:
        inferred_deal = _first_active_deal_for_contact(case.related_contact)
        if inferred_deal:
            case.deal = inferred_deal
            changed_fields.append("deal")
    if case.lead:
        normalized_lead_name = f"{case.lead.first_name} {case.lead.last_name}".strip()
        if not case.lead_name:
            case.lead_name = normalized_lead_name
            changed_fields.append("lead_name")
        if not case.lead_source and case.lead.lead_source:
            case.lead_source = case.lead.lead_source
            changed_fields.append("lead_source")
        if not case.email and case.lead.email:
            case.email = case.lead.email
            changed_fields.append("email")
        if not case.phone and (case.lead.phone or case.lead.mobile):
            case.phone = case.lead.phone or case.lead.mobile
            changed_fields.append("phone")
        if not case.company and case.lead.company:
            case.company = case.lead.company
            changed_fields.append("company")
        if not case.reported_by:
            case.reported_by = normalized_lead_name
            changed_fields.append("reported_by")
    if not case.case_reason:
        case.case_reason = _default_case_reason(case.type)
        changed_fields.append("case_reason")
    if not case.subject:
        case.subject = _case_subject_for(
            {
                "product": case.product,
                "related_contact": case.related_contact,
            }
        )
        changed_fields.append("subject")
    case.updated_by = user
    if changed_fields:
        changed_fields.append("updated_by")
        case.save(update_fields=changed_fields + ["updated_at"])
        _sync_support_links(
            "case",
            case,
            contact=case.related_contact,
            account=case.account,
            deal=case.deal,
            product=case.product,
            lead=case.lead,
        )
        log_timeline(
            "case",
            case.pk,
            "updated",
            "Case Updated",
            user=user,
            metadata={"updated_fields": changed_fields},
        )
    return case


@transaction.atomic
def delete_case(case: SupportCase, user):
    case.is_active = False
    case.updated_by = user
    case.save(update_fields=["is_active", "updated_by", "updated_at"])
    log_timeline("case", case.pk, "deleted", "Case Deleted", user=user)


@transaction.atomic
def create_solution(data: dict[str, Any], user):
    payload = data.copy()
    payload.setdefault("owner", user)
    payload.setdefault("status", "Draft")
    source_case = payload.get("source_case")
    if source_case:
        payload.setdefault("product", source_case.product)
        payload.setdefault("question", source_case.subject)
        existing_solution = (
            SupportSolution.objects.filter(is_active=True, source_case=source_case)
            .order_by("-created_at")
            .first()
        )
        if existing_solution:
            payload.pop("created_by", None)
            payload.pop("updated_by", None)
            return update_solution(existing_solution, payload, user)
    payload.setdefault("solution_title", _solution_title_for(payload))
    payload["created_by"] = user
    payload["updated_by"] = user
    solution = SupportSolution.objects.create(**payload)
    _sync_support_links(
        "solution",
        solution,
        contact=source_case.related_contact if source_case else None,
        account=source_case.account if source_case else None,
        deal=source_case.deal if source_case else None,
        product=solution.product,
        case=source_case,
        lead=source_case.lead if source_case else None,
    )
    log_timeline("solution", solution.pk, "created", "Solution Created", user=user, metadata={"solution_number": solution.solution_number})
    return solution


@transaction.atomic
def update_solution(solution: SupportSolution, data: dict[str, Any], user):
    changed_fields = []
    for field, value in data.items():
        if getattr(solution, field) != value:
            setattr(solution, field, value)
            changed_fields.append(field)
    if solution.source_case and not solution.product:
        solution.product = solution.source_case.product
        changed_fields.append("product")
    if solution.source_case and not solution.question:
        solution.question = solution.source_case.subject
        changed_fields.append("question")
    if not solution.solution_title:
        solution.solution_title = _solution_title_for(
            {"source_case": solution.source_case, "product": solution.product}
        )
        changed_fields.append("solution_title")
    solution.updated_by = user
    if changed_fields:
        changed_fields.append("updated_by")
        solution.save(update_fields=changed_fields + ["updated_at"])
        _sync_support_links(
            "solution",
            solution,
            contact=solution.source_case.related_contact if solution.source_case else None,
            account=solution.source_case.account if solution.source_case else None,
            deal=solution.source_case.deal if solution.source_case else None,
            product=solution.product,
            case=solution.source_case,
            lead=solution.source_case.lead if solution.source_case else None,
        )
        log_timeline(
            "solution",
            solution.pk,
            "updated",
            "Solution Updated",
            user=user,
            metadata={"updated_fields": changed_fields},
        )
    return solution


@transaction.atomic
def delete_solution(solution: SupportSolution, user):
    solution.is_active = False
    solution.updated_by = user
    solution.save(update_fields=["is_active", "updated_by", "updated_at"])
    log_timeline("solution", solution.pk, "deleted", "Solution Deleted", user=user)


def _record_kwargs(module_type: str, obj):
    return {"case": obj} if module_type == "case" else {"solution": obj}


@transaction.atomic
def create_note(module_type: str, obj, note: str, user):
    note_obj = SupportNote.objects.create(**_record_kwargs(module_type, obj), note=note, created_by=user)
    log_timeline(module_type, obj.pk, "note_added", "Note added", user=user)
    return note_obj


@transaction.atomic
def create_comment(module_type: str, obj, comment: str, user):
    comment_obj = SupportComment.objects.create(**_record_kwargs(module_type, obj), comment=comment, created_by=user)
    if module_type == "case":
        SupportCase.objects.filter(pk=obj.pk).update(no_of_comments=models.F("no_of_comments") + 1)
        obj.refresh_from_db(fields=["no_of_comments"])
    else:
        SupportSolution.objects.filter(pk=obj.pk).update(no_of_comments=models.F("no_of_comments") + 1)
        obj.refresh_from_db(fields=["no_of_comments"])
    log_timeline(module_type, obj.pk, "comment_added", "Comment added", user=user)
    return comment_obj


@transaction.atomic
def add_attachment(module_type: str, obj, file, user):
    attachment = SupportAttachment.objects.create(
        **_record_kwargs(module_type, obj),
        file=file,
        original_name=getattr(file, "name", "attachment"),
        file_type=getattr(file, "content_type", Path(getattr(file, "name", "")).suffix.lstrip(".")),
        uploaded_by=user,
    )
    log_timeline(module_type, obj.pk, "attachment_added", "Attachment added", user=user, metadata={"file_name": attachment.original_name})
    return attachment


@transaction.atomic
def add_linked_record(module_type: str, obj, data: dict[str, Any], user):
    linked = SupportLinkedRecord.objects.create(**_record_kwargs(module_type, obj), **data)
    log_timeline(module_type, obj.pk, "record_linked", "Connected record linked", user=user, metadata={"relationship_label": linked.relationship_label})
    return linked


@transaction.atomic
def add_activity(module_type: str, obj, data: dict[str, Any], user):
    activity = SupportActivity.objects.create(**_record_kwargs(module_type, obj), created_by=user, **data)
    log_timeline(module_type, obj.pk, "activity_added", data.get("action") or "Activity added", user=user)
    return activity


@transaction.atomic
def add_email_log(module_type: str, obj, data: dict[str, Any], user):
    email_log = SupportEmailLog.objects.create(**_record_kwargs(module_type, obj), sent_by=user, **data)
    log_timeline(module_type, obj.pk, "email_logged", "Email logged", user=user, metadata={"subject": email_log.subject})
    return email_log


def create_import_job(module_type: str, file, uploaded_by, operation: str | None = None, duplicate_check_field: str | None = None):
    file.seek(0)
    headers, rows = _parse_uploaded_file(file)
    file.seek(0)
    job = SupportImportJob.objects.create(
        module_type=module_type,
        file=file,
        original_name=file.name,
        file_type=Path(file.name).suffix.lower().lstrip("."),
        operation=operation or "add",
        duplicate_check_field=duplicate_check_field or "",
        uploaded_by=uploaded_by,
        headers=headers,
        sample_rows=rows[:5],
    )
    return job


def get_import_job(job_id: int, module_type: str):
    return SupportImportJob.objects.get(pk=job_id, module_type=module_type, is_active=True)


def inspect_import_job(job: SupportImportJob):
    required_fields = ["subject"] if job.module_type == "case" else ["solution_title", "question", "answer"]
    duplicate_options = ["case_id", "subject", "email"] if job.module_type == "case" else ["solution_id", "solution_title"]
    field_map = CASE_IMPORT_FIELD_MAP if job.module_type == "case" else SOLUTION_IMPORT_FIELD_MAP
    suggestions = {header: field_map.get(header.strip().lower(), "") for header in job.headers}
    duplicate_targets = {}
    used_targets = set()
    warnings = []
    for source, target in suggestions.items():
        if not target:
            continue
        if target in used_targets:
            warnings.append(f"CRM field '{target}' is mapped more than once.")
        used_targets.add(target)
        duplicate_targets[source] = target
    job.status = SupportImportJob.Status.VALIDATED
    job.save(update_fields=["status", "updated_at"])
    return {
        "job_id": job.id,
        "module_type": job.module_type,
        "status": job.status,
        "required_fields": required_fields,
        "duplicate_check_options": duplicate_options,
        "suggested_mapping": duplicate_targets,
        "warnings": warnings,
        "sample_rows": job.sample_rows,
        "headers": job.headers,
    }


def _normalize_import_mapping(module_type: str, mapping: dict[str, str]):
    field_map = CASE_IMPORT_FIELD_MAP if module_type == "case" else SOLUTION_IMPORT_FIELD_MAP
    normalized = {}
    duplicate_targets = set()
    warnings = []
    for source, target in mapping.items():
        normalized_target = field_map.get((target or "").strip().lower(), target)
        if normalized_target in normalized.values():
            duplicate_targets.add(normalized_target)
        normalized[source] = normalized_target
    for target in sorted(duplicate_targets):
        warnings.append(f"CRM field '{target}' is mapped more than once.")
    return normalized, warnings


def _required_import_fields(module_type: str):
    return ["subject"] if module_type == "case" else ["solution_title", "question", "answer"]


def _duplicate_options(module_type: str):
    return ["case_number", "subject", "email"] if module_type == "case" else ["solution_number", "solution_title"]


def _load_import_rows(job: SupportImportJob):
    file_handle = job.file.open("rb")
    try:
        headers, rows = _parse_uploaded_file(file_handle)
    finally:
        file_handle.close()
    return headers, rows


def _apply_defaults(payload: dict[str, Any], defaults: dict[str, Any]):
    for key, value in defaults.items():
        if payload.get(key) in (None, ""):
            payload[key] = value
    return payload


def _resolve_import_payload(module_type: str, row: dict[str, Any], mapping: dict[str, str], defaults: dict[str, Any], user):
    payload: dict[str, Any] = {}
    connected_value = None
    for source, target in mapping.items():
        if not target:
            continue
        raw_value = _clean_value(row.get(source))
        if raw_value is None:
            continue
        if target == "connected_to":
            connected_value = raw_value
            continue
        if target in {"owner", "product", "related_contact", "account", "deal"}:
            payload[target] = _resolve_fk(target, raw_value)
        else:
            payload[target] = raw_value
    payload = _apply_defaults(payload, defaults)
    if module_type == "case":
        payload.setdefault("owner", user)
    else:
        payload.setdefault("owner", user)
    return payload, connected_value


def _find_existing_record(module_type: str, payload: dict[str, Any], duplicate_check_field: str | None):
    if module_type == "case":
        queryset = SupportCase.objects.filter(is_active=True)
    else:
        queryset = SupportSolution.objects.filter(is_active=True)
    if not duplicate_check_field:
        return None
    value = payload.get(duplicate_check_field)
    if value in (None, ""):
        return None
    lookup = {f"{duplicate_check_field}__iexact": str(value)} if isinstance(value, str) else {duplicate_check_field: value}
    return queryset.filter(**lookup).first()


@transaction.atomic
def execute_import_job(
    module_type: str,
    job: SupportImportJob,
    field_mapping: dict[str, str],
    default_values: dict[str, Any],
    operation: str,
    duplicate_check_field: str | None,
    automation_enabled: bool,
    user,
):
    normalized_mapping, warnings = _normalize_import_mapping(module_type, field_mapping)
    required_fields = _required_import_fields(module_type)
    mapped_targets = set(target for target in normalized_mapping.values() if target)
    missing_required = [field for field in required_fields if field not in mapped_targets and field not in default_values]
    if missing_required:
        raise DjangoValidationError(f"Mandatory fields missing from mapping: {', '.join(missing_required)}")

    _, rows = _load_import_rows(job)
    errors = []
    imported_count = 0
    updated_count = 0
    skipped_count = 0

    for index, row in enumerate(rows, start=1):
        payload, connected_to = _resolve_import_payload(module_type, row, normalized_mapping, default_values, user)
        missing_row_fields = [field for field in required_fields if not payload.get(field)]
        if missing_row_fields:
            errors.append({"row": index, "errors": {field: ["This field is required."] for field in missing_row_fields}})
            skipped_count += 1
            continue

        duplicate_field = duplicate_check_field or ("subject" if module_type == "case" else "solution_title")
        existing = _find_existing_record(module_type, payload, duplicate_field)

        try:
            if existing:
                if operation == "add":
                    skipped_count += 1
                    errors.append({"row": index, "errors": {duplicate_field: ["Duplicate record exists."]}})
                    continue
                if module_type == "case":
                    update_case(existing, payload, user)
                    record = existing
                else:
                    update_solution(existing, payload, user)
                    record = existing
                updated_count += 1
                log_timeline(module_type, record.pk, "import_updated", f"{record.__class__.__name__.replace('Support', '')} updated from import", user=user, metadata={"automation_enabled": automation_enabled})
            else:
                if operation == "update":
                    skipped_count += 1
                    errors.append({"row": index, "errors": {duplicate_field: ["No matching record found for update."]}})
                    continue
                if module_type == "case":
                    record = create_case(payload, user)
                else:
                    record = create_solution(payload, user)
                imported_count += 1
                log_timeline(module_type, record.pk, "import_created", f"{record.__class__.__name__.replace('Support', '')} created from import", user=user, metadata={"automation_enabled": automation_enabled})

            if connected_to and module_type == "solution":
                linked_data = _resolve_connected_to_value(str(connected_to))
                if linked_data:
                    add_linked_record(module_type, record, linked_data, user)
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": index, "errors": {"non_field_errors": [str(exc)]}})
            skipped_count += 1

    job.operation = operation
    job.duplicate_check_field = duplicate_check_field or ""
    job.field_mapping = normalized_mapping
    job.default_values = default_values
    job.automation_enabled = automation_enabled
    job.validation_errors = errors + [{"warnings": warnings}] if warnings else errors
    job.imported_count = imported_count
    job.updated_count = updated_count
    job.skipped_count = skipped_count
    job.error_count = len(errors)
    job.result_summary = {
        "total_rows": len(rows),
        "warnings": warnings,
        "automation_enabled": automation_enabled,
    }
    job.status = SupportImportJob.Status.COMPLETED if not errors else SupportImportJob.Status.FAILED
    job.save(
        update_fields=[
            "operation",
            "duplicate_check_field",
            "field_mapping",
            "default_values",
            "automation_enabled",
            "validation_errors",
            "imported_count",
            "updated_count",
            "skipped_count",
            "error_count",
            "result_summary",
            "status",
            "updated_at",
        ]
    )
    return job


def _resolve_connected_to_value(value: str):
    value = value.strip()
    if not value:
        return None
    account = Account.objects.filter(account_name__iexact=value, is_active=True).first()
    if account:
        return {"account": account, "relationship_label": "Connected To"}
    contact = Contact.objects.filter(
        Q(first_name__iexact=value) | Q(last_name__iexact=value) | Q(email__iexact=value),
        is_active=True,
    ).first()
    if contact:
        return {"contact": contact, "relationship_label": "Connected To"}
    deal = Deal.objects.filter(deal_name__iexact=value, is_active=True).first()
    if deal:
        return {"deal": deal, "relationship_label": "Connected To"}
    product = Product.objects.filter(Q(product_name__iexact=value) | Q(product_code__iexact=value), is_active=True).first()
    if product:
        return {"product": product, "relationship_label": "Connected To"}
    vendor = Vendor.objects.filter(vendor_name__iexact=value, is_active=True).first()
    if vendor:
        return {"vendor": vendor, "relationship_label": "Connected To"}
    lead = Lead.objects.filter(
        Q(first_name__iexact=value) | Q(last_name__iexact=value) | Q(email__iexact=value)
    ).first()
    if lead:
        return {"lead": lead, "relationship_label": "Connected To"}
    return None


def build_lookup_payload(lookup_name: str, query: str = ""):
    term = query.strip()
    if lookup_name == "products":
        queryset = Product.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(Q(product_name__icontains=term) | Q(product_code__icontains=term))
        return [
            {
                "id": obj.id,
                "name": obj.product_name,
                "label": f"{obj.product_name} ({obj.product_code})" if obj.product_code else obj.product_name,
                "product_code": obj.product_code,
                "unit_price": obj.unit_price,
            }
            for obj in queryset[:25]
        ]
    if lookup_name == "accounts":
        queryset = Account.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(account_name__icontains=term)
        return [{"id": obj.id, "name": obj.account_name, "label": obj.account_name} for obj in queryset[:25]]
    if lookup_name == "contacts":
        queryset = Contact.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(
                Q(first_name__icontains=term) | Q(last_name__icontains=term) | Q(email__icontains=term)
            )
        return [
            {
                "id": obj.id,
                "name": f"{obj.first_name} {obj.last_name}".strip(),
                "label": f"{obj.first_name} {obj.last_name}".strip(),
                "email": obj.email,
                "phone": obj.phone,
                "account_id": obj.account_id,
                "account_name": obj.account.account_name if obj.account else None,
            }
            for obj in queryset[:25]
        ]
    if lookup_name == "vendors":
        queryset = Vendor.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(Q(vendor_name__icontains=term) | Q(email__icontains=term))
        return [{"id": obj.id, "name": obj.vendor_name, "label": obj.vendor_name, "email": obj.email, "phone": obj.phone} for obj in queryset[:25]]
    if lookup_name == "deals":
        queryset = Deal.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(deal_name__icontains=term)
        return [{"id": obj.id, "name": obj.deal_name, "label": obj.deal_name} for obj in queryset[:25]]
    if lookup_name == "cases":
        queryset = SupportCase.objects.filter(is_active=True)
        if term:
            queryset = queryset.filter(Q(case_number__icontains=term) | Q(subject__icontains=term))
        return [
            {
                "id": obj.id,
                "name": obj.case_number or obj.subject,
                "label": obj.case_number or obj.subject,
                "subject": obj.subject,
                "status": obj.status,
            }
            for obj in queryset[:25]
        ]
    if lookup_name == "leads":
        queryset = Lead.objects.all()
        if term:
            queryset = queryset.filter(
                Q(first_name__icontains=term)
                | Q(last_name__icontains=term)
                | Q(email__icontains=term)
                | Q(company__icontains=term)
            )
        return [
            {
                "id": obj.id,
                "name": f"{obj.first_name} {obj.last_name}".strip(),
                "label": f"{obj.first_name} {obj.last_name}".strip() or obj.email,
                "email": obj.email,
                "phone": obj.phone or obj.mobile,
                "account_name": obj.company,
                "lead_source": obj.lead_source,
            }
            for obj in queryset[:25]
        ]
    raise DjangoValidationError("Lookup not found.")


def quick_create_product(data: dict[str, Any], user):
    payload = data.copy()
    payload.setdefault("owner", user)
    return Product.objects.create(**payload)
